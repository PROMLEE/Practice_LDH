import random
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Color, Border, Side, Alignment

def AutoFitColumnSize(worksheet, columns=None, margin=2):
    for i, column_cells in enumerate(worksheet.columns):
        is_ok = False
        if columns == None:
            is_ok = True
        elif isinstance(columns, list) and i in columns:
            is_ok = True

        if is_ok:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = length + margin
    return worksheet


pages = []
voca = []
Vocafile = openpyxl.load_workbook('VOCA_BIBLE.xlsx')
Testfile = openpyxl.load_workbook('test_voca.xlsx')
print('시험지를 만들 day를 입력하세요.')
a = 0
for i in Vocafile.sheetnames:
    print('%6s'%i, end = '')
    a+=1
    if a%5 == 0:
        print()
day = input('\nex) 1-5,19\n')
day = day.split(',')
for i in day:
    if '-' in i:
        a = list(map(int, i.split('-')))
        b = list(range(a[0], a[1]+1))
        pages += b
    else:
        pages.append(int(i))
pages = set(pages)
pages = list(pages)
pages.sort()
for i in pages:
    sheet_name = 'day' + str(i)
    sheet = Vocafile[sheet_name]
    for row in sheet.iter_rows(min_row=1, max_row=25, min_col=1, max_col=2):
        a = []
        for cell in row:
            a.append(cell.value)
        voca.append(a)
random.shuffle(voca)
ws = Testfile.get_sheet_by_name('test')
Testfile.remove(ws)
Testfile.create_sheet('test')
test_sheet = Testfile['test']
row_num = 1
col_num = 1
w_font = Font(name='calibri', size=20)
w_font2 = Font(name='HY견고딕', size=7, color="F2F2F2")
THIN_BORDER = Border(Side('thin'), Side('thin'), Side('thin'), Side('thin'))
for i in range(len(voca)):
    test_sheet.cell(row=row_num, column=col_num).value = voca[i][0]
    test_sheet.cell(row=row_num, column=col_num).font = w_font
    test_sheet.cell(row=row_num, column=col_num).border = THIN_BORDER
    col_num += 1
    test_sheet.cell(row=row_num, column=col_num).value = '%10s' % ' '
    test_sheet.cell(row=row_num, column=col_num).border = THIN_BORDER
    col_num += 1
    test_sheet.cell(row=row_num, column=col_num).value = voca[i][1]
    test_sheet.cell(row=row_num, column=col_num).font = w_font2
    test_sheet.cell(row=row_num, column=col_num).border = THIN_BORDER
    test_sheet.cell(row=row_num, column=col_num).alignment = Alignment(horizontal='left')
    col_num -= 2
    row_num += 1
    if row_num > 25:
        row_num = 1
        col_num += 3
AutoFitColumnSize(test_sheet, None, 10)
# test_sheet.print_area = 'A1:C25','D1:F25'
test_sheet.oddFooter.left.text = "Page &[Page] of &N"
test_sheet.oddFooter.left.size = 10
test_sheet.oddFooter.left.font = "Tahoma,Bold"
Testfile.save('test_voca.xlsx')

print('day',pages,' 의 단어 시험지가 완성되었습니다. (test_voca.xlsx)')